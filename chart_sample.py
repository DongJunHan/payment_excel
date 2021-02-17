from openpyxl import chart
from openpyxl import load_workbook
from openpyxl import Workbook

path = "./시가총액 Top 15.xlsx"
def make_chart(path):
    wb = load_workbook(path,data_only=True)
    ws = wb.active
    chart_val = chart.LineChart()
    chart_val.title = "주가_일별_등락패턴"
    chart_val.x_axis.title = "date"
    chart_val.y_axis.title = "stock"

    #data영역임.
    datas = chart.Reference(ws,min_col = 1,min_row = 2, max_col = ws.max_column, max_row = ws.max_row)
    chart_val.add_data(datas,from_rows=True,titles_from_data=True)
    #카테고리영역임. min_col,min_row는 시작범위 cell번호이고, max_col,max_row는 최대 범위 cell번호 이다. min_col = 2이므로 2열부터 시작이고, min_row = 1라서 1행부터 시작, max_col = 21 21열까지이고, max_row = 1 1행까지임.
    cats = chart.Reference(ws,min_col=2, min_row = 1, max_col = 21, max_row = 1)
    chart_val.set_categories(cats)
    #어느 cell에 차트를 그릴건지 정함.
    ws.add_chart(chart_val,"A17")
    wb.save("./차트결과.xlsx")
    wb.close()

make_chart(path)