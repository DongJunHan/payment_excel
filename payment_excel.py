from openpyxl import load_workbook
from operator import itemgetter

path = "./한동준_국민체크8906.xlsx"
writePath = "./OOO_법인카드지출내역서.xlsx"
def get_excel_data(path):
    #시트 로드. data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
    try:
        loadWb = load_workbook(path,data_only=True)
    except FileNotFoundError:
        print("get_excel_data There is no file")
        return

    #시트 이름으로 불러오기
    loadWs = loadWb['Sheet1']

    #cell의 데이터 가져오기.
    paymentList =[]
    for row in loadWs.rows:
        rowList = []
        for cell in row:
            rowList.append(cell.value)
        paymentList.append(rowList)
    return paymentList

def parsing_excel_data(dataList):
    cellList = []
    for data in dataList:
        date = data[0]
        if date != "승인일":
            cell = []
            cell.append(data[0]) # date
            cell.append(data[1]) # time 내용을 점심식사, 저녁식사구분하기 위함.
            cell.append(data[6]) # shop name
            cell.append(data[7]) # shop 식대인지, 택시인지 구분하기 위함. TODO. 현재는 식대랑, 택시만 사용할거같지만 추후에 어떤것이 더 생길지 모름
            cell.append(data[10]) # price
            cellList.append(cell)
    return cellList

def write_excel(path,dataList):
    redataList = sorted(dataList, key=itemgetter(0))
    print(redataList)
    try:
        loadWb = load_workbook(path,data_only=False)
        sheet = loadWb.get_sheet_by_name('고정비')
        # sheet = loadWb.get_active_sheet()
        rowNum = 6
        totalAmount = 0
        for row in redataList:
            columnNum = 2
            #승인일자
            sheet.cell(row = rowNum, column = columnNum, value = row[0])
            columnNum += 1
            #상호명
            sheet.cell(row = rowNum, column = columnNum, value = row[2])
            columnNum += 1
            #금액
            sheet.cell(row = rowNum, column = columnNum, value = row[4])
            totalAmount += int(row[4])
            columnNum += 1
            #업무
            if row[3] == "택시":
                sheet.cell(row = rowNum, column = columnNum, value = "OmniDoc")
                columnNum += 1
                #내용
                sheet.cell(row = rowNum, column = columnNum, value = row[3])
            else:
                sheet.cell(row = rowNum, column = columnNum, value = "공통비")
                columnNum += 1
                #내용
                if (row[1][0:2].find('11') != -1) or (row[1][0:2].find('12') != -1) or (row[1][0:2].find('13') != -1) or (row[1][0:2].find('14') != -1):
                    sheet.cell(row = rowNum, column = columnNum, value = "점심식사")
                elif (row[1][0:2].find('17') != -1) or (row[1][0:2].find('18') != -1) or (row[1][0:2].find('19') != -1) or (row[1][0:2].find('20') != -1):
                    sheet.cell(row = rowNum, column = columnNum, value = "저녁식사")
                else:
                    sheet.cell(row = rowNum, column = columnNum, value = "식대")

            

            # for data in row:
            #     sheet.cell(row = rowNum, column = columnNum, value = data)
            #     columnNum += 1
            rowNum += 1
        sheet.cell(row = 3, column = 4, value = totalAmount)
    except FileNotFoundError:
        print("write_excel There is no file")
        return
    loadWb.save(path)

dataList = get_excel_data(path)
parseData = parsing_excel_data(dataList)
# print(parseData)
write_excel(writePath,parseData)